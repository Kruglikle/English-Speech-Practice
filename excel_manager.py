from openpyxl import load_workbook
from typing import List, Tuple
import random

class ExcelManager:
    def __init__(self, file_path: str = "tenderisthenight.xlsx"):
        self.file_path = file_path
        self.wb = load_workbook(filename=self.file_path)
        self.sheet = self.wb.active


    def get_all_pairs(self) -> List[Tuple[str, str]]:
        rows = list(self.sheet.iter_rows(min_row=2, max_col=2, values_only=True))
        filtered = [
            (eng, rus) for eng, rus in rows
            if eng and eng.strip() != "" and rus and rus.strip() != ""
        ]
        return filtered

    def get_random_pair(self) -> Tuple[str, str]:
        """Возвращает случайную пару (английская фраза, русский перевод), без пустых"""
        pairs = self.get_all_pairs()
        if not pairs:
            raise ValueError("Файл пуст или все строки пустые!")
        return random.choice(pairs)

    def get_random_words(self, n: int = 10) -> List[str]:
        """Возвращает случайные английские слова"""
        all_english_words = [row[0] for row in self.get_all_pairs()]
        return random.sample(all_english_words, min(len(all_english_words), n))

    def add_phrase(self, english: str, russian: str):
        """Добавляет новую фразу в файл"""
        self.sheet.append([english, russian])
        self.wb.save(self.file_path)
